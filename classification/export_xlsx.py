"""
classification/export_xlsx.py

Part 2 - Step 4c: export the deliverable table as an XLSX spreadsheet
(submitted on moo.uni1.de).

Columns (exactly as specified on PDF page 28):
    repository_id
    project_type
    project_title
    primary_class
    secondary_class      // if any
    no_project_files     // number of files in the project in total

One row per project. Written to deliverables/23455702-sq26-classification-table.xlsx

Run:
    python -m classification.export_xlsx
"""

import sys
import pathlib
import sqlite3

_root = pathlib.Path(__file__).resolve().parent.parent
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

DEFAULT_DB = str(_root / "23455702-sq26-classification.db")
OUT_DIR = _root / "deliverables"
STUDENT_ID = "23455702"

COLUMNS = ["repository_id", "project_type", "project_title",
           "primary_class", "secondary_class", "no_project_files"]


def export_xlsx(db_path: str = DEFAULT_DB) -> pathlib.Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    OUT_DIR.mkdir(exist_ok=True)
    conn = sqlite3.connect(db_path)
    rows = conn.execute(
        """SELECT repository_id, type, title, primary_class,
                  secondary_class, no_project_files
           FROM projects
           ORDER BY repository_id,
                    CASE type WHEN 'QDA_PROJECT' THEN 0 WHEN 'QD_PROJECT' THEN 1
                              WHEN 'OTHER_PROJECT' THEN 2 ELSE 3 END,
                    id"""
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "classification"

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(COLUMNS)
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for r in rows:
        ws.append(list(r))

    # column widths
    widths = [14, 15, 60, 42, 42, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows)+1}"

    out = OUT_DIR / f"{STUDENT_ID}-sq26-classification-table.xlsx"
    wb.save(out)
    print(f"[XLSX] Wrote {len(rows)} rows -> {out}")
    return out


if __name__ == "__main__":
    db = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_DB
    export_xlsx(db)
